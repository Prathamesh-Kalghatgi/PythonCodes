from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook("C:\\Users\\kalgh\\OneDrive\\Desktop\\Friends.xlsx")

# Extract data from sheet1
sheet1 = workbook['Sheet1']
data_sheet1 = []

for row in sheet1.iter_rows():
    data_row = []
    for cell in row:
        data_row.append(cell.value)
    data_sheet1.append(data_row)

# Extract data from sheet2
sheet2 = workbook['Sheet2']
data_sheet2 = []

for row in sheet2.iter_rows():
    data_row = []
    for cell in row:
        data_row.append(cell.value)
    data_sheet2.append(data_row)

# printing the data from sheet1
print("Data from Sheet1:")
for row in data_sheet1:
    print(row)

# Printing data from Sheet2
print("Data from Sheet2:")
for row in data_sheet2:
    print(row)
